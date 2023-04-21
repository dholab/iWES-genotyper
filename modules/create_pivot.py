#!/usr/bin/env python
from argparse import ArgumentParser, ArgumentTypeError
import pandas as pd
import numpy as np
import json
import math
import os
from openpyxl.styles import Font, Fill, PatternFill, GradientFill
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from scipy import stats
from Bio import SeqIO


def str2bool(v):
    if isinstance(v, bool):
        return v
    if v.lower() in ('yes', 'true', 't', 'y', '1'):
        return True
    elif v.lower() in ('no', 'false', 'f', 'n', '0'):
        return False
    else:
        raise ArgumentTypeError('Boolean value expected.')


def parse_process_arrays_args(parser: ArgumentParser):
    """Parses the python script arguments from bash and makes sure files/inputs are valid"""
    parser.add_argument('--out_dir',
                        type=str,
                        help='directory to output the results',
                        required=True)
    parser.add_argument('--project_name',
                        type=str,
                        help='Project name for file prefix',
                        required=True)
    parser.add_argument('--ref_dir',
                        type=str,
                        help='where the config files are stored',
                        default=None,
                        required=False)
    parser.add_argument('--animal_lookup_path',
                        type=str,
                        help='2 column csv file each row is a file prefix to corresponding sample id',
                        required=True)
    parser.add_argument('--haplotype_lookup',
                        type=str,
                        help='your haplotype look up created with the program to create the summerized haplotypes from \
                        create_ref_fasta_lookups.py',
                        default=None,
                        required=False)
    parser.add_argument('--confidence_coeff_json',
                        type=str,
                        help='json with confidence coeff created before. This file should be updated as verified data \
                        is added to the database',
                        default=None,
                        required=False)
    parser.add_argument('--conf_group',
                    type=str,
                    help='which haplotype group to use, or use NO_FILTER if you do not want to choose a group',
                    default='NO_FILTER',
                    required=False)

    parser.add_argument('--depth_threshold',
                        type=int,
                        default=3,
                        required=False)
    parser.add_argument('-e',
                        '--db_ext',
                        nargs='+',
                        default=None,
                        help='extensions to add at end of the headers to descriminate files (i.e. -e gen exon miseq',
                        required=False)


def get_process_arrays_args():
    """	Inputs arguments from bash
    Gets the arguments, checks requirements, returns a dictionary of arguments
    Return: args - Arguments as a dictionary
    """
    parser = ArgumentParser()
    parse_process_arrays_args(parser)
    args = parser.parse_args()
    return args


#################################
# Read in the arguments         #
# Details are in the help above #
#################################
args = get_process_arrays_args()
# same arguments to a local variable by same name as the argument
out_dir = args.out_dir
project_name = args.project_name
ref_dir = args.ref_dir
haplotype_lookup = args.haplotype_lookup
animal_lookup_path = args.animal_lookup_path
confidence_coeff_json = args.confidence_coeff_json
depth_threshold = args.depth_threshold
db_ext = args.db_ext
conf_group = args.conf_group
################################################################
# Replace non declared files with relative paths to ref_dir #
################################################################

if haplotype_lookup is None:
    haplotype_lookup = os.path.join(ref_dir, 'haplotype_lookup.csv')

if confidence_coeff_json is None:
    confidence_coeff_json = os.path.join(ref_dir, 'confidence_coeff.json')


xlsx_filepath = None
pd.options.mode.chained_assignment = None

def error_code(x, y, z):
    '''
    Converts haplotypes to proper formatting
    1 is homozygous (add a dash)
    2 is return each haplotype
    >2 is a concatenated TMH of each haplotype possibility
    :param x: Haplotype for that cell
    :param y: Number of potential haplotypes called
    :param z: Concatenated haplotypes of potential call
    :return: Return the list if it is homozygous (Haplo, -), it is a row based function
    THat later gets exploded and deduplicated per method.
    '''
    if y > 2:
        return 'TMH:({0})'.format(z)
    if y == 2:
        return x
    if y == 1:
        # add a dash for a second option, as it needs to be looked at to some degree
        return [x, '-']
    return "NO HAPLO"


def fasta_to_df(fasta_path=None, header_name='allele', sequence_name='SEQUENCE', as_df=False):
    fasta_sequences = SeqIO.parse(open(fasta_path), 'fasta')
    fasta_dict = {}
    for fasta in fasta_sequences:
        name, sequence = fasta.id, str(fasta.seq)
        fasta_dict[name] = sequence
    fasta_sequences.close()
    if as_df:
        return pd.DataFrame(fasta_dict.items(), columns=[header_name, sequence_name])
    return fasta_dict


def get_confidence(x, y, predict_dict, threshold):
    '''
    This is based on a pre-staged linear regression performed in the median data vs the minimum data for called alleles
    :param x: the median value for the species
    :param y: the species/dicitionary key
    :param predict_dict: dictionary of species(method) vs a list of regression [coeeff, intercept and std dev]
    :param threshold: The minimum depth of coverage for a position to be called
    :return: the one-sided p-score, which is the expected percentage of alleles called
    '''
    # get the stats of interest
    coeff = predict_dict[y][0]
    intercept = predict_dict[y][1]
    stdev = predict_dict[y][2]
    # calculate z-score
    z_score = ((coeff * x + intercept) - threshold) / stdev
    return round(1 - stats.norm.sf(abs(z_score)), 4)


#####################
# Open Lookup files #
# open all of the look up files
lookup_dicts = {}
df_lu_dict = {}
ambig_dict_temp = {}
for i in range(1, len(db_ext)):
    db_i = db_ext[i]
    df_lu_dict[db_i] = pd.DataFrame()
    suffix_list = db_ext[0:i]
    for db_c_i in suffix_list:
        print(db_ext[i], db_c_i)
        lookup_json_path = os.path.join(ref_dir, '{0}_to_{1}_lookup.json'.format(db_i, db_c_i))
        with open(lookup_json_path) as f_in:
            lookup_dict_i = json.load(f_in)
        for key, value in lookup_dict_i.items():
            df_temp = pd.DataFrame({'{0}_allele'.format(db_i): key, 'allele': value})
            df_lu_dict[db_i] = pd.concat([df_lu_dict[db_i], df_temp], ignore_index=True)
            if key in ambig_dict_temp.keys():
                ambig_dict_temp[key] = ambig_dict_temp[key] + value
            else:
                ambig_dict_temp[key] = value
for i in range(1, len(db_ext)):
    if os.path.exists(os.path.join(ref_dir, '{0}_lu.csv'.format(db_ext[i]))):
        df_exon_lu = pd.read_csv(os.path.join(ref_dir, 'exon_lu.csv'))
        exon_lu_dict = {}
        for index, row in df_exon_lu.iterrows():
            if row['allele'] in ambig_dict_temp.keys():
                ambig_dict_temp[row['allele']] = ambig_dict_temp[row['allele']] + [row['allele_grouped'].split('|')[1]]
            else:
                ambig_dict_temp[row['allele']] = [row['allele_grouped'].split('|')[1]]
ambig_dict = {}
for diag, ipd_list in ambig_dict_temp.items():
    ipd_list = ['-'.join(x.split('-')[1:]) for x in ipd_list]
    diag = '-'.join(diag.split('-')[1:])
    ambig_dict[diag] = ','.join(ipd_list)

                # df_lu = pd.concat([df_lu, pd.DataFrame({'db_allele': key, 'allele': value})], ignore_index=True)
# Extract dictionaries of need #

try:
    with open(confidence_coeff_json) as f_in:
        confidence_coeff_dict = json.load(f_in)
except:
    confidence_coeff_dict = {}
#  End open and parse look up files #
#####################################
# Create  a output directory if it does not exist #
os.makedirs(out_dir, exist_ok=True)
# Open the Data file of medians
df_norm_median = pd.read_csv(os.path.join(out_dir, '{0}_norm_median_all.csv'.format(project_name)))
# The sample is often a number, let coerce it to be a string.
df_norm_median['SAMPLE_NUM'] = df_norm_median['SAMPLE_NUM'].astype(str)
df_norm_median.rename(columns={'ALLELE': 'allele',
                                 'SAMPLE_NUM': 'accession',
                                 'DEPTH_ADJ': 'read_ct'}, inplace=True)
# The open the read count of aligned reads per sample file
df_read_ct = pd.read_csv(os.path.join(out_dir, '{0}_read_ct.csv'.format(project_name)))
# gs_id is the SAMPLE_NUM coerce it to be the same
# gs_id is the format the GS teams have been reporting as an internal ID for a sample, vs a customer id
# a customer ids may not be unqiue and we use an internal unique identifier
df_read_ct['gs_id'] = df_read_ct['gs_id'].astype(str)

df_norm_median = df_norm_median[df_norm_median['MAX_GAP'] < 120]
# Filter for the genomic data of the median file
df_dict = {}

if len(db_ext) < 2:
    df = df_norm_median
else:
    df = df_norm_median[df_norm_median['DB'] == db_ext[0]]

df.sort_values(by=['accession', 'allele'], inplace=True)
for i in range(1, len(db_ext)):
    print(db_ext[i])
    db_i = db_ext[i]
    suffix_list = db_ext[0:i]
    # for db_c_i in suffix_list:
    if len(df_lu_dict[db_i]) > 0:
        # extract the median db of choice
        df_m_i = df_norm_median[df_norm_median['DB'] == db_i]
        df_m_i.rename(columns={'allele': '{0}_allele'.format(db_i),
                                      'SAMPLE_NUM': 'accession',
                                      'DEPTH_ADJ': 'read_ct',
                                      'unique_maps_per_allele': 'unique_maps_per_allele_{0}'.format(db_i)},
                             inplace=True)
        df_m_i = df_m_i[
            (df_m_i['Depth_Threshold_From_End_Position'] == 0) & (df_m_i['Depth_Threshold_From_Start_Position'] == 0)]
        df_m_i = df_m_i.merge(df_lu_dict[db_i], on=['{0}_allele'.format(db_i)], how='left')
        df_m_i['allele'] = df_m_i['allele'].astype(str)
        df_norm_median_i = df_norm_median[df_norm_median['DB'].isin(suffix_list)]
        df_norm_median_i = df_norm_median_i[['allele', 'accession', 'unique_maps_per_allele']]
        df_m_i = df_m_i.merge(df_norm_median_i, on=['allele', 'accession'], how='left')
        # fill what didn't match with 0 to denote a non-match instead of a na to ease with later aggregations
        df_m_i = df_m_i.fillna(0)
        # group them by them mapping to figure out which have maps and which dont
        df_m_i = df_m_i.groupby(['{0}_allele'.format(db_i), 'read_ct', 'accession', 'unique_maps_per_allele_{0}'.format(db_i)])[
            'unique_maps_per_allele'].max().reset_index().rename(columns={'{0}_allele'.format(db_i): 'allele'})
        # If the count is zero, it means the miseq call is not a subset of any genomic/exon2 call and needs to be retained
        df_m_i = df_m_i[df_m_i['unique_maps_per_allele'] == 0]
        df_m_i.drop(columns=['unique_maps_per_allele'], inplace=True)
        df_m_i.rename(columns={'unique_maps_per_allele_miseq': 'unique_maps_per_allele'}, inplace=True)
        df_m_i.sort_values(by=['accession', 'allele'], inplace=True)
        df = pd.concat([df, df_m_i], ignore_index=True)
    else:
        df = pd.concat([df, df_norm_median[df_norm_median['DB'] == db_i]], ignore_index=True)

df_i_f = df.copy()

#################################################################
# End retain the miseq and df_exon calls based on the heirarchy #
#################################################################
##############################################################
# Start Pivot the read count data so it can latter be joined #
##############################################################
df_read_ct_pivot = df_read_ct.pivot_table(values='sample_read_ct',
                                          columns=['gs_id'],
                                          aggfunc=np.max,
                                          fill_value=0).reset_index()
df_read_ct_pivot.rename_axis(['index2'], inplace=True, axis=1)
df_read_ct_pivot.rename(columns={'index': 'gs_id'}, inplace=True)
df_read_ct_pivot.rename_axis(['index'], inplace=True, axis=1)
######################################################
# End the read count data so it can latter be joined #
######################################################
##############################################################################
# Determine and pivot median counts/per sample  to help determine over calls #
##############################################################################
df_median_counts = df.groupby(['accession'])['read_ct'].median().reset_index().rename(
    columns={'read_ct': 'normalized_median_allele_count'})
df_median_counts['normalized_median_allele_count'] = df_median_counts['normalized_median_allele_count'].round(0)
# Report blank as a filler for confidence interval if it has not been calculated yet.
if len(confidence_coeff_dict):
    df_median_counts['P_VALUE'] = [get_confidence(x,
                                                  conf_group,
                                                  confidence_coeff_dict,
                                                  depth_threshold) for x in
                                   df_median_counts['normalized_median_allele_count']]
else:
    df_median_counts['P_VALUE'] = ''
df_median_counts_pivot = df_median_counts.pivot_table(values=['normalized_median_allele_count', 'P_VALUE'],
                                                      columns=['accession'],
                                                      aggfunc=np.max,
                                                      fill_value=0).reset_index()

df_median_counts_pivot.rename_axis(['index2'], inplace=True, axis=1)
df_median_counts_pivot.rename(columns={'index': 'gs_id'}, inplace=True)
df_median_counts_pivot.rename_axis(['index'], inplace=True, axis=1)

# rename the df from accesion to gs_id
df.rename(columns={'accession': 'gs_id'}, inplace=True)

# copy the df here as we need this data to work directly with the haplotype calls
df_haplo_ready = df.copy()
df_haplo_ready.rename(columns={'allele_ipd': 'allele'}, inplace=True)
df_haplo_ready = df_haplo_ready[['allele', 'read_ct', 'gs_id']]

# join animal data (if it exists)
# Warning: This will filter for things only on the list so make sure every column has a corresponding entry
print(animal_lookup_path, os.path.exists(animal_lookup_path))
if os.path.exists(animal_lookup_path):
    animal_lookup = pd.read_csv(animal_lookup_path)
    animal_lookup = animal_lookup[['gs_id']]
    animal_lookup_dict = {}
    animal_lookup['gs_id'] = animal_lookup['gs_id'].astype(str)
    df = df.merge(animal_lookup[['gs_id']], how='inner', on=['gs_id'])
    df_read_ct = df_read_ct.merge(animal_lookup[['gs_id']], how='inner', on=['gs_id'])

# print('__{}__'.format(species))
#
    # Remove the Mamu- because it is redundant and takes space

# df['MHC_TYPE'] = [x.split('_')[1] if int((y * 100) % 100) in [49, 99] else x for x, y in
#                   zip(df['allele'], df['read_ct'])]
# df['TYPE'] = [type_dict[x] if x in type_dict.keys() else x for x in df['MHC_TYPE']]
# Find what group the allele is in for the dynamic aggregation
df.to_csv(os.path.join(out_dir, 'concat_haplotype.csv'), index=False)

xlsx_filepath = os.path.join(out_dir, '{0}.pivot.xlsx'.format(project_name))
df['# Obs'] = df.groupby('allele')['read_ct'].transform('count')
# change to get normalized values
gs_id_list = list(df['gs_id'].unique())
gs_id_list.sort()
df_gs_id = pd.DataFrame([gs_id_list], columns=gs_id_list)

genotype_pivot = df.pivot_table(values='read_ct',
                                index=['allele', '# Obs'],
                                columns=['gs_id'],
                                aggfunc=np.max,
                                fill_value='').reset_index()

genotype_pivot.rename_axis(['index'], inplace=True, axis=1)
genotype_pivot.rename(columns={'allele': 'gs_id'}, inplace=True)
print('df_read_ct_pivot')

df_allele_count_summary = df.groupby('gs_id')['read_ct'].count().reset_index()
df_allele_count_summary.rename(columns={'read_ct': '# Alleles Identified'}, inplace=True)
df_count_summary_pivot = df_allele_count_summary.pivot_table(values='# Alleles Identified',
                                                             columns=['gs_id'],
                                                             aggfunc=np.max,
                                                             fill_value=0).reset_index()
df_count_summary_pivot.rename_axis(['index2'], inplace=True, axis=1)
df_count_summary_pivot.rename(columns={'index': 'gs_id'}, inplace=True)
df_count_summary_pivot.rename_axis(['index'], inplace=True, axis=1)

genotype_pivot['header_name'] = genotype_pivot['gs_id']
genotype_pivot['gs_id'] = ['-'.join(x.split('-')[1:]) if x.count('-') > 1 else x for x in genotype_pivot['gs_id']]

genotype_pivot['ambiguous_alleles'] = [x if x not in ambig_dict.keys() else ambig_dict[x] for x in
                                       genotype_pivot['gs_id']]


genotype_pivot['gs_id'] = genotype_pivot['gs_id'].str.replace('__', '*')
genotype_pivot['gs_id'] = genotype_pivot['gs_id'].str.replace('_', ':')
genotype_pivot['gs_id'] = genotype_pivot['gs_id'].str.replace('-diag', '-miseq')
genotype_pivot['gs_id'] = genotype_pivot['gs_id'].str.replace('-nuc', '-cdna')
genotype_pivot['gs_id'] = genotype_pivot['gs_id'].str.replace('A1:028g1-miseq', 'AG3:02g1_A028-miseq')
genotype_pivot['gs_id'] = genotype_pivot['gs_id'].str.replace('A1:110g1-miseq', 'A1:110g1-cdna-miseq')
genotype_pivot['gs_id'] = genotype_pivot['gs_id'].str.replace('A1:119g1-miseq', 'AG3:02g2_A119-miseq')
genotype_pivot['gs_id'] = [x.replace(':', '*', 1) for x in genotype_pivot['gs_id']]
genotype_pivot['gs_id'] = genotype_pivot['gs_id'].str.replace('-{0}'.format(db_ext[0]), '')
genotype_pivot.sort_values(by=['gs_id'], inplace=True)
df_list = [df_gs_id, df_count_summary_pivot, df_read_ct_pivot, df_median_counts_pivot, genotype_pivot]
# create the header for the pivot table
df_haplo_pivot = pd.DataFrame()
################################
# Algorithm to call haplotypes #
################################

if os.path.exists(haplotype_lookup):
    # open the haplotype_lookup csv
    df_haplo_call = pd.read_csv(haplotype_lookup)
    df_haplo_ready.rename(columns={'allele': 'bait_allele'}, inplace=True)
    # merge the lookup to each call
    df_haplo_merge = df_haplo_call.merge(df_haplo_ready[['bait_allele', 'gs_id']], on=['bait_allele'], how='inner')


    df_haplo_merge = df_haplo_merge[
        ['HAPLOTYPE_CALL', 'PREFIX', 'TYPE', 'allele', 'gs_id', 'CALL_COUNT']].drop_duplicates()
    # find the counts of each call to figure out if all the allele types are covered
    df_haplo_merge['CALL_GSID_COUNT'] = df_haplo_merge.groupby(['HAPLOTYPE_CALL', 'PREFIX', 'TYPE', 'gs_id'])[
        'allele'].transform('count')
    # retain only the hapotypes that are fully covered (all alleles are present to make tha hpalotype call)
    df_haplo_pass = df_haplo_merge[df_haplo_merge['CALL_COUNT'] == df_haplo_merge['CALL_GSID_COUNT']]
    # drop the duplicates and remove the allele names, we just care about the haplotype call
    df_haplo_pass = df_haplo_pass[['HAPLOTYPE_CALL', 'PREFIX', 'TYPE', 'gs_id', 'CALL_COUNT']].drop_duplicates()
    # figure out how many haplotypes are call per group (1: Homozygous, 2: Heterozygous, >2 = Too many Haplotypes)
    df_haplo_pass['PASS_COUNT'] = df_haplo_pass.groupby(['gs_id', 'TYPE'])['HAPLOTYPE_CALL'].transform('count')
    # rank them so we can retain the proper ones in the right order.
    df_haplo_pass['rank'] = df_haplo_pass.groupby(['gs_id', 'TYPE'])['CALL_COUNT'].transform('rank', method='first')
    # join to create a list for the TMH
    df_haplo_pass['TMH'] = df_haplo_pass.groupby(['gs_id', 'TYPE'])['HAPLOTYPE_CALL'].transform('; '.join)
    # Now use t he error code to report TMH or Heterozygous.
    df_haplo_pass['HAPLOTYPE_FINAL'] = [error_code(x, y, z) for x, y, z in
                                        zip(df_haplo_pass['HAPLOTYPE_CALL'], df_haplo_pass['PASS_COUNT'],
                                            df_haplo_pass['TMH'])]
    # if it has only one haplotype, it will make a list [haplotype_call, '-'] which then can be xploded
    # Explode the list to yield a 1 and 2 haplotype based on the rankings
    df_haplo_pass = df_haplo_pass.explode('HAPLOTYPE_FINAL')
    # Rank again to as stuff may have been filtered
    df_haplo_pass['rank'] = df_haplo_pass.groupby(['gs_id', 'TYPE'])['CALL_COUNT'].transform('rank', method='first')
    # Get the type for each haplotype (Group 1 and Group 2) and match the format of the haplotype_to_header pivot_list
    df_haplo_pass['TYPE'] = ['{0} {1}'.format(x.replace('_', ' '), int(y)) for x, y in
                             zip(df_haplo_pass['TYPE'], df_haplo_pass['rank'])]
    # figure out what only report the first 2 haplotypes as too many haplotypes may have 3 and it would be redundant
    df_haplo_pass = df_haplo_pass[df_haplo_pass['rank'] < 3]
    # any thing that is blank (No called haplotypes) will be a fill value of No Haplotype
    # This is exploiting a agg function that currently works with a fill value of strings
    df_haplo_pivot = df_haplo_pass.pivot_table(values='HAPLOTYPE_FINAL',
                                               index=['TYPE'],
                                               columns=['gs_id'],
                                               aggfunc=np.max,
                                               fill_value='NO HAPLO').reset_index()
    if len(df_haplo_pivot) < 1:
        df_haplo_pivot = pd.DataFrame()
    else:
        # Get the indexes correct
        df_haplo_pivot.rename_axis(['index'], inplace=True, axis=1)
        # Get the Column names aligned with the other pivot tables to be concatenated
        df_haplo_pivot.rename(columns={'TYPE': 'gs_id'}, inplace=True)

df_xlx_pivot = pd.concat(
    [df_gs_id, df_count_summary_pivot, df_haplo_pivot, df_read_ct_pivot, df_median_counts_pivot, genotype_pivot],
    ignore_index=True)

df_xlx_pivot.rename(columns={'gs_id': 'Animal IDs'}, inplace=True)
# concatenate if there ia valid animal look up path as the column names will be different
sample_list = []
if (len(animal_lookup_path) > 0) and os.path.exists(animal_lookup_path):
    animal_lookup = pd.read_csv(animal_lookup_path)
    animal_lookup_dict = {}
    animal_lookup['gs_id'] = animal_lookup['gs_id'].astype(str)
    sample_list = list(animal_lookup['gs_id'])
    for idx, row in animal_lookup.iterrows():
        animal_lookup_dict[row['gs_id']] = row['animal_id']
    gs_id_list = list(df_xlx_pivot.columns)
    df_xlx_pivot.rename(columns=animal_lookup_dict, inplace=True)
    animal_list = list(animal_lookup['animal_id'])
    animal_list2 = []
    for gs_id_i in gs_id_list:
        if gs_id_i in animal_lookup_dict.keys():
            animal_list2.append(animal_lookup_dict[gs_id_i])
    animal_list2.sort()
    col_order = ['Animal IDs', '# Obs'] + animal_list2 + ['header_name', 'ambiguous_alleles']
    # concatenate if there ia not a valid animal look up path
else:
    col_names = list(df_xlx_pivot.columns)
    col_names.remove('Animal IDs')
    col_names.remove('# Obs')
    col_names.remove('ambiguous_alleles')
    col_names.remove('header_name')
    col_order = ['Animal IDs', '# Obs'] + col_names + ['header_name', 'ambiguous_alleles']
    animal_list2 = col_names

# reorder the columns
df_xlx_pivot = df_xlx_pivot[col_order]
first_item = True
first_list = list(df_xlx_pivot['Animal IDs'])
skip_rows = first_list.index('normalized_median_allele_count')
df_xlx_pivot.to_excel(xlsx_filepath, sheet_name='Sheetname_1', index=False)

##################################################
# This is additional highlighting and formatting #
##################################################
# create a key for the highlighting
wb = load_workbook(filename=xlsx_filepath)
ws = wb.worksheets[0]


def get_maximum_cols(sheet):
    max_col = 0
    for i in range(1, 20000):
        max_col += 1
        if sheet.cell(row=1, column=i).value == None:
            break
    return max_col


max_column = get_maximum_cols(ws)
color_key_col = get_column_letter(max_column + 1)
desc_key_col = get_column_letter(max_column + 2)

cell_1 = ws['{0}2'.format(color_key_col)]
cell_1.value = 'Black-Bold Text'
cell_1.font = Font(bold=True)
cell_1 = ws['{0}3'.format(color_key_col)]
cell_1.value = 'Purple Highlight'
cell_1.fill = PatternFill(start_color="F2CAFA", fill_type="solid")
cell_1 = ws['{0}4'.format(color_key_col)]
cell_1.value = 'Green Text'
cell_1.font = Font(color='006400')
cell_1 = ws['{0}5'.format(color_key_col)]
cell_1.value = 'Red Text'
cell_1.font = Font(color='ad0a02')
cell_1 = ws['{0}6'.format(color_key_col)]
cell_1.value = 'Yellow Background'
cell_1.fill = PatternFill(start_color="F2DFA9", fill_type="solid")
cell_1 = ws['{0}2'.format(desc_key_col)]
cell_1.value = "Fully mapped genomic allele and There is at least 1 read that uniquely maps to the allele and not any other viable allele"
cell_1 = ws['{0}3'.format(desc_key_col)]
cell_1.value = "Fully mapped genomic allele and There is not any reads that uniquely map to the allele and not any other viable allele"
cell_1 = ws['{0}4'.format(desc_key_col)]
cell_1.value = "Fully mapped miseq db or exon2  allele and There is at least 1 read that uniquely maps to the allele and not any other viable allele"
cell_1 = ws['{0}5'.format(desc_key_col)]
cell_1.value = "Fully mapped miseq db or exon2  allele and There is not any reads that uniquely map to the allele and not any other viable allele"
cell_1 = ws['{0}6'.format(desc_key_col)]
cell_1.value = "The normalized median count is less than  3/4 f the expected normalized median. This means there is a likely overcall."
#####################################################
# Add comments depth from edge and max gap comments #
#####################################################
allelle_list = []
row_start = skip_rows + 3
col_start = 3

column_list = []
for i in range(col_start, 20000):
    if ws.cell(row=2, column=i).value == None:
        max_col = i
        break
    column_list.append(ws.cell(row=2, column=i).value)

header_list = []
allelle_list = []
for i in range(col_start, 20000):
    if ws.cell(row=1, column=i).value == None:
        max_col = i
        break
    header_list.append(ws.cell(row=1, column=i).value)

col_index = header_list.index('header_name') + col_start
for i in range(row_start, 20000):
    if ws.cell(row=i, column=col_index).value == None:
        max_col = i
        break
    allelle_list.append(ws.cell(row=i, column=col_index).value)
if len(sample_list) > 0:
    df_i_f = df_i_f[df_i_f['accession'].isin(sample_list)]
print(df_i_f)
print(allelle_list)
for idx, row in df_i_f.iterrows():
    column_i = column_list.index(str(row['accession'])) + col_start

    col_letter_i = get_column_letter(column_i)
    if row['allele'] in allelle_list:
        row_i = allelle_list.index(row['allele']) + row_start
        comment_i_list = []
        attribute_list = []
        # light purple: F2CAFA
        # dark green: #006400
        # yellow: F2DFA9
        # dark purple: 9900FF
        if ws['{0}{1}'.format(col_letter_i, row_i)].value is None:
            continue
        if row['unique_maps_per_allele'] == 1:
            ws['{0}{1}'.format(col_letter_i, row_i)].font = Font(bold=True)
            if not row['allele'].endswith('-{0}'.format(db_ext[0])):
                ws['{0}{1}'.format(col_letter_i, row_i)].font = Font(color="006400", bold=True)
            if row['DEPTH_NORM'] < .75:
                # attribute_list.append('background-color: #F2DFA9')
                ws['{0}{1}'.format(col_letter_i, row_i)].fill = PatternFill(start_color="F2DFA9", fill_type="solid")
        else:
            if row['allele'].endswith('-{0}'.format(db_ext[0])):
                if row['DEPTH_NORM'] < .75:
                    # attribute_list.append('background-color: #F2DFA9')
                    ws['{0}{1}'.format(col_letter_i, row_i)].fill = GradientFill(stop=("F2DFA9", 'F2CAFA'))
                else:
                    ws['{0}{1}'.format(col_letter_i, row_i)].fill = PatternFill(start_color="F2CAFA", fill_type="solid")
                    # ws['{0}{1}'.format(col_letter_i, row_i)].font = Font(color="9900FF")
            else:
                if row['DEPTH_NORM'] < .75:
                    # attribute_list.append('background-color: #F2DFA9')
                    ws['{0}{1}'.format(col_letter_i, row_i)].fill = GradientFill(stop=("F2DFA9", 'FF0000'))
                else:
                    ws['{0}{1}'.format(col_letter_i, row_i)].fill = PatternFill(start_color="FF0000", fill_type="solid")
                    # ws['{0}{1}'.format(col_letter_i, row_i)].font = Font(color="FF0000")
        if row['Depth_Threshold_From_Start_Position'] > 0:
            comment_i_list.append(
                'Depth_Threshold_From_Start_Position: {0}'.format(row['Depth_Threshold_From_Start_Position']))
        if row['Depth_Threshold_From_End_Position'] > 0:
            comment_i_list.append(
                'Depth_Threshold_From_End_Position: {0}'.format(row['Depth_Threshold_From_End_Position']))
        if row['MAX_GAP'] > 70:
            comment_i_list.append('MAX_GAP: {0}, Gap_Position: {1}'.format(row['MAX_GAP'], row['GAP_POSITION']))
        if len(comment_i_list) > 0:
            comment_i = '\n'.join(comment_i_list)
            ws['{0}{1}'.format(col_letter_i, row_i)].comment = Comment(comment_i, 'z')
mycell = ws['C{0}'.format(skip_rows + 3)]
ws.freeze_panes = mycell
wb.save(filename=xlsx_filepath)
